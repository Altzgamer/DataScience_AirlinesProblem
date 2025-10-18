import csv
import sqlite3
import xml.etree.ElementTree as ET
from pathlib import Path
import time
import logging
import re
import os
import json
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from ruamel.yaml import YAML
from ruamel.yaml.error import YAMLError

PROCESS_CSV = False
PROCESS_TAB = False
PROCESS_XML = False
PROCESS_YAML = False
PROCESS_PDF = False #
PROCESS_JSON = True
PROCESS_XLS = True

CLEAR_CSV = False
CLEAR_TAB = False
CLEAR_XML = False
CLEAR_YAML = False
CLEAR_PDF = False
CLEAR_JSON = False
CLEAR_XLS = False

CSV_FILE = 'Data/BoardingData.csv'
TAB_FILE = 'Data/Sirena-export-fixed.tab'
XML_FILE = 'Data/PointzAggregator-AirlinesData.xml'
YAML_FILE = 'Data/SkyTeam-Exchange.yaml'
PDF_FILE = 'Data/Skyteam_Timetable.pdf'
JSON_FILE = 'Data/FrequentFlyerForum-Profiles.json'
XLS_DIR = 'Data/YourBoardingPassDotAero'
DB_FILE = 'DataBase.db'

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def create_database_connection(db_file: str) -> tuple[sqlite3.Connection, sqlite3.Cursor]:
    """Create a connection to the SQLite database and return the connection and cursor."""
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        return conn, cursor
    except sqlite3.Error as e:
        logger.error(f"Database connection error: {e}")
        raise

def create_boarding_data_table(cursor: sqlite3.Cursor) -> None:
    """Create the boarding_data table if it doesn't exist."""
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS boarding_data (
            PassengerFirstName TEXT,
            PassengerSecondName TEXT,
            PassengerLastName TEXT,
            PassengerSex TEXT,
            PassengerBirthDate TEXT,
            PassengerDocument TEXT,
            BookingCode TEXT,
            TicketNumber TEXT,
            Baggage TEXT,
            FlightDate TEXT,
            FlightTime TEXT,
            FlightNumber TEXT,
            CodeShare TEXT,
            Destination TEXT
        )
    ''')

def create_sirena_data_table(cursor: sqlite3.Cursor) -> None:
    """Create the sirena_data table if it doesn't exist."""
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sirena_data (
            PaxName TEXT,
            PaxBirthDate TEXT,
            DepartDate TEXT,
            DepartTime TEXT,
            ArrivalDate TEXT,
            ArrivalTime TEXT,
            FlightCode TEXT,
            FromAirport TEXT,
            Dest TEXT,
            Code TEXT,
            e_Ticket TEXT,
            TravelDoc TEXT,
            Seat TEXT,
            Meal TEXT,
            TrvCls TEXT,
            Fare TEXT,
            Baggage TEXT,
            PaxAdditionalInfo TEXT,
            AgentInfo TEXT
        )
    ''')

def create_pointz_aggregator_table(cursor: sqlite3.Cursor) -> None:
    """Create the pointz_aggregator_data table if it doesn't exist."""
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pointz_aggregator_data (
            UserUID TEXT,
            FirstName TEXT,
            LastName TEXT,
            CardNumber TEXT,
            BonusProgramm TEXT,
            FlightCode TEXT,
            FlightDate TEXT,
            Departure TEXT,
            Arrival TEXT,
            Fare TEXT
        )
    ''')

def create_skyteam_data_table(cursor: sqlite3.Cursor) -> None:
    """Create the skyteam_data table if it doesn't exist."""
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS skyteam_data (
            FlightDate TEXT,
            FlightNumber TEXT,
            FFProgram TEXT,
            FFNumber TEXT,
            TravelClass TEXT,
            Fare TEXT,
            Departure TEXT,
            Arrival TEXT,
            Status TEXT
        )
    ''')

def create_skyteam_timetable_table(cursor: sqlite3.Cursor) -> None:
    """Create the skyteam_timetable table if it doesn't exist."""
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS skyteam_timetable (
            from_city TEXT,
            from_country TEXT,
            from_code TEXT,
            to_city TEXT,
            to_country TEXT,
            to_code TEXT,
            validity TEXT,
            days TEXT,
            dep_time TEXT,
            arr_time TEXT,
            flight TEXT,
            aircraft TEXT,
            travel_time TEXT
        )
    ''')

def create_frequent_flyer_profiles_table(cursor: sqlite3.Cursor) -> None:
    """Create the frequent_flyer_profiles table if it doesn't exist."""
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS frequent_flyer_profiles (
            Nick TEXT PRIMARY KEY,
            Sex TEXT,
            FirstName TEXT,
            LastName TEXT,
            TravelDocuments TEXT,
            Loyalties TEXT
        )
    ''')

def create_frequent_flyer_flights_table(cursor: sqlite3.Cursor) -> None:
    """Create the frequent_flyer_flights table if it doesn't exist."""
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS frequent_flyer_flights (
            NickName TEXT,
            FlightDate TEXT,
            Flight TEXT,
            Codeshare INTEGER,
            DepartureCity TEXT,
            DepartureAirport TEXT,
            DepartureCountry TEXT,
            ArrivalCity TEXT,
            ArrivalAirport TEXT,
            ArrivalCountry TEXT,
            FOREIGN KEY (NickName) REFERENCES frequent_flyer_profiles(NickName)
        )
    ''')

def create_boarding_pass_xls_table(cursor: sqlite3.Cursor) -> None:
    """Create the boarding_pass_xls table if it doesn't exist."""
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS boarding_pass_xls (
            PassengerTitle TEXT,
            PassengerName TEXT,
            LoyaltyProgram TEXT,
            LoyaltyNumber TEXT,
            FareClass TEXT,
            FlightNumber TEXT,
            DepartureCity TEXT,
            ArrivalCity TEXT,
            DepartureAirport TEXT,
            ArrivalAirport TEXT,
            FlightDate TEXT,
            FlightTime TEXT,
            PNR TEXT,
            ETicket TEXT
        )
    ''')

def parse_csv_file(cursor: sqlite3.Cursor, csv_file: str) -> None:
    """Parse the CSV file and insert data into the boarding_data table."""
    try:
        with open(csv_file, 'r', encoding='utf-8') as file:
            reader = csv.reader(file, delimiter=';')
            next(reader)  # Skip header
            for row in reader:
                if len(row) == 14:  # Validate row length
                    cursor.execute('''
                        INSERT INTO boarding_data (
                            PassengerFirstName, PassengerSecondName, PassengerLastName, PassengerSex,
                            PassengerBirthDate, PassengerDocument, BookingCode, TicketNumber,
                            Baggage, FlightDate, FlightTime, FlightNumber, CodeShare, Destination
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', row)
                else:
                    logger.warning(f"Skipping invalid CSV row: {row}")
    except FileNotFoundError:
        logger.error(f"CSV file not found: {csv_file}")
        raise
    except Exception as e:
        logger.error(f"Error parsing CSV file: {e}")
        raise

def parse_tab_file(cursor: sqlite3.Cursor, tab_file: str) -> None:
    """Parse the TAB file and insert data into the sirena_data table."""
    try:
        with open(tab_file, 'r', encoding='utf-8') as file:
            next(file)  # Skip header
            for line in file:
                line = line.rstrip()
                if not line:
                    continue
                try:
                    row = [
                        line[0:60].strip(),   # PaxName
                        line[60:72].strip(),  # PaxBirthDate
                        line[72:84].strip(),  # DepartDate
                        line[84:96].strip(),  # DepartTime
                        line[96:108].strip(), # ArrivalDate
                        line[108:120].strip(),# ArrivalTime
                        line[120:132].strip(),# FlightCode
                        line[132:138].strip(),# FromAirport
                        line[138:144].strip(),# Dest
                        line[144:150].strip(),# Code
                        line[150:168].strip(),# e_Ticket
                        line[168:180].strip(),# TravelDoc
                        line[180:186].strip(),# Seat
                        line[186:192].strip(),# Meal
                        line[192:198].strip(),# TrvCls
                        line[198:216].strip(),# Fare
                        line[216:240].strip(),# Baggage
                        line[240:276].strip(),# PaxAdditionalInfo
                        line[276:336].strip() # AgentInfo
                    ]
                    if len(row) == 19:
                        cursor.execute('''
                            INSERT INTO sirena_data (
                                PaxName, PaxBirthDate, DepartDate, DepartTime, ArrivalDate,
                                ArrivalTime, FlightCode, FromAirport, Dest, Code,
                                e_Ticket, TravelDoc, Seat, Meal, TrvCls,
                                Fare, Baggage, PaxAdditionalInfo, AgentInfo
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', row)
                    else:
                        logger.warning(f"Skipping invalid TAB row: {row}")
                except IndexError:
                    logger.warning(f"Skipping malformed TAB line: {line}")
    except FileNotFoundError:
        logger.error(f"TAB file not found: {tab_file}")
        raise
    except Exception as e:
        logger.error(f"Error parsing TAB file: {e}")
        raise

def parse_xml_file(cursor: sqlite3.Cursor, xml_file: str) -> None:
    """Parse the XML file and insert data into the pointz_aggregator_data table."""
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()

        for user in root.findall('user'):
            uid = user.get('uid', '')
            name = user.find('name')
            first_name = name.get('first', '') if name is not None else ''
            last_name = name.get('last', '') if name is not None else ''

            cards = user.find('cards')
            if cards is not None:
                for card in cards.findall('card'):
                    card_number = card.get('number', '')
                    bonus_programm = card.find('bonusprogramm').text if card.find('bonusprogramm') is not None else ''

                    activities = card.find('activities')
                    if activities is not None:
                        for activity in activities.findall('activity'):
                            if activity.get('type') == 'Flight':
                                flight_code = activity.find('Code').text if activity.find('Code') is not None else ''
                                flight_date = activity.find('Date').text if activity.find('Date') is not None else ''
                                departure = activity.find('Departure').text if activity.find('Departure') is not None else ''
                                arrival = activity.find('Arrival').text if activity.find('Arrival') is not None else ''
                                fare = activity.find('Fare').text if activity.find('Fare') is not None else ''

                                cursor.execute('''
                                    INSERT INTO pointz_aggregator_data (
                                        UserUID, FirstName, LastName, CardNumber, BonusProgramm,
                                        FlightCode, FlightDate, Departure, Arrival, Fare
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ''', (uid, first_name, last_name, card_number, bonus_programm,
                                      flight_code, flight_date, departure, arrival, fare))
    except FileNotFoundError:
        logger.error(f"XML file not found: {xml_file}")
        raise
    except ET.ParseError:
        logger.error(f"Invalid XML format in file: {xml_file}")
        raise
    except Exception as e:
        logger.error(f"Error parsing XML file: {e}")
        raise

def parse_yaml_file(cursor: sqlite3.Cursor, yaml_file: str) -> None:
    """Parse the YAML file and insert data into the skyteam_data table."""
    try:
        yaml = YAML(typ='safe')  # Use ruamel.yaml with safe loading
        with open(yaml_file, 'r', encoding='utf-8') as file:
            start_time = time.time()
            logger.info(f"Starting YAML parsing for {yaml_file}")
            data = yaml.load(file)
            if not data:
                logger.warning(f"YAML file {yaml_file} is empty")
                return

            rows = []
            for flight_date, flights in data.items():
                for flight_number, flight_info in flights.items():
                    departure = flight_info.get('FROM', '')
                    arrival = flight_info.get('TO', '')
                    status = flight_info.get('STATUS', '')
                    ff_data = flight_info.get('FF', {})

                    for ff_number, ff_info in ff_data.items():
                        ff_program = ff_number.split()[0] if ff_number and ' ' in ff_number else ''
                        travel_class = ff_info.get('CLASS', '')
                        fare = ff_info.get('FARE', '')

                        rows.append((
                            flight_date, flight_number, ff_program, ff_number,
                            travel_class, fare, departure, arrival, status
                        ))

            if rows:
                cursor.executemany('''
                    INSERT INTO skyteam_data (
                        FlightDate, FlightNumber, FFProgram, FFNumber, 
                        TravelClass, Fare, Departure, Arrival, Status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', rows)

            elapsed = time.time() - start_time
            logger.info(f"Completed YAML parsing in {elapsed:.2f} seconds")
    except FileNotFoundError:
        logger.error(f"YAML file not found: {yaml_file}")
        raise
    except YAMLError as e:
        logger.error(f"Invalid YAML format in file {yaml_file}: {e}")
        raise
    except Exception as e:
        logger.error(f"Error parsing YAML file: {e}")
        raise

def process_pdf_to_excel(pdf_file, excel_file, start_from_page):
    def extract_tables_from_pdf(pdf_path, excel_path, start_page=1):
        all_tables = []
        current_parts = []
        current_title = None

        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                if page_num < start_page:
                    continue

                print(f"Обрабатывается страница {page_num}")

                text = page.extract_text()

                table_titles = re.findall(r'FROM', text)

                tables = page.extract_tables()

                for i, table in enumerate(tables):
                    if not table or not any(
                            any(cell is not None and str(cell).strip() for cell in row) for row in table):
                        continue

                    df = pd.DataFrame(table)

                    df = df.dropna(how='all').dropna(axis=1, how='all')

                    if df.empty:
                        continue

                    table_title = None
                    if i < len(table_titles):
                        table_title = table_titles[i]
                    elif table_titles:
                        table_title = table_titles[0]

                    if table_title:
                        if current_parts:
                            all_tables.append((current_title, current_parts))
                            current_parts = []

                        current_title = table_title
                        current_parts = [df]
                    else:
                        if current_parts:
                            current_parts.append(df)
                        else:
                            current_title = f"Таблица со страницы {page_num}"
                            current_parts = [df]

        if current_parts:
            all_tables.append((current_title, current_parts))

        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Sheet1'
        sheet2 = wb.create_sheet('Sheet2')

        row_offset = 1
        row_offset2 = 1
        for title, parts in all_tables:
            if title:
                sheet.cell(row=row_offset, column=1, value=title)
                row_offset += 2

            if not parts:
                continue

            main_df = parts[0]

            row_offset += 1

            # Убираем жирный шрифт и границы для первой строки
            for r_idx in range(len(main_df)):
                row = main_df.iloc[r_idx]
                for c, value in enumerate(row, start=1):
                    cell = None
                    if c < 10:
                        cell = sheet.cell(row=row_offset, column=c,
                                          value=str(value).strip() if value is not None else '')
                    else:
                        cell = sheet2.cell(row=row_offset, column=c - 9,
                                           value=str(value).strip() if value is not None else '')

                    # Убираем жирный шрифт и границы для первой строки
                    if r_idx == 0:  # первая строка
                        if cell:
                            cell.font = Font(bold=False)  # Убираем жирный текст
                            cell.border = Border()  # Убираем границы
                row_offset += 1
                row_offset2 += 1

            shift_after = [1, 6, 10, 15]

            for cont_df in parts[1:]:
                for r_idx in range(len(cont_df)):
                    row = cont_df.iloc[r_idx]
                    excel_col = 1
                    for orig_idx, value in enumerate(row):
                        orig_1based = orig_idx + 1
                        cell = None
                        if orig_1based < 10:
                            cell = sheet.cell(row=row_offset, column=excel_col,
                                              value=str(value).strip() if value is not None else '')
                        else:
                            cell = sheet2.cell(row=row_offset, column=excel_col - 9,
                                               value=str(value).strip() if value is not None else '')
                        excel_col += 1
                        if orig_1based in shift_after:
                            excel_col += 1
                        # Убираем жирный шрифт и границы для первой строки
                        if r_idx == 0:  # первая строка
                            if cell:
                                cell.font = Font(bold=False)
                                cell.border = Border()
                    row_offset += 1

            row_offset += 2

        wb.save(excel_path)

    def remove_first_row_from_xlsx(input_file, output_file):
        workbook = openpyxl.load_workbook(input_file)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            if worksheet.max_row <= 1:
                print(f"Лист '{sheet_name}' имеет 0 или 1 строку — пропускаем.")
                continue

            rows_to_shift = []
            for row_num in range(2, worksheet.max_row + 1):
                row_data = []
                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    row_data.append(cell.value)
                rows_to_shift.append(row_data)

            # Очищаем все данные на листе
            for row_num in range(1, worksheet.max_row + 1):
                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = None

            # Записываем сдвинутые строки начиная с первой
            for new_row_num, row_data in enumerate(rows_to_shift, start=1):
                for col_num, cell_value in enumerate(row_data, start=1):
                    worksheet.cell(row=new_row_num, column=col_num).value = cell_value

            # Делаем первую строку не жирным и без границ
            first_row = worksheet[1]
            for cell in first_row:
                cell.font = Font(bold=False)  # Убираем жирный текст
                cell.border = Border()  # Убираем границы

        workbook.save(output_file)
        print(f"Новый файл сохранён как '{output_file}'.")

    def split_tables(df):
        empty_mask = df.isnull().all(axis=1)

        non_empty_groups = []
        start = 0
        for i in range(len(df)):
            if empty_mask.iloc[i]:
                if i > start:
                    non_empty_groups.append((start, i))
                start = i + 1
        if len(df) > start:
            non_empty_groups.append((start, len(df)))

        tables = []
        for s, e in non_empty_groups:
            if e - s >= 2:
                table = df.iloc[s:e].reset_index(drop=True)
                table = table.dropna(how='all').reset_index(drop=True)
                if len(table) >= 2:
                    tables.append(table)
        return tables

    extract_tables_from_pdf(pdf_file, excel_file, start_from_page)

    df1 = pd.read_excel(excel_file, sheet_name='Sheet1')
    tables1 = split_tables(df1)

    df2_full = pd.read_excel(excel_file, sheet_name='Sheet2')
    df2 = df2_full.iloc[:, 2:]
    tables2 = split_tables(df2)

    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        for i, table in enumerate(tables1):
            sheet_name = f'Sheet1_Table_{i + 1}'
            table.to_excel(writer, sheet_name=sheet_name, index=False)

        for i, table in enumerate(tables2):
            sheet_name = f'Sheet2_Table_{i + 1}'
            table.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Processed and saved {len(tables1) + len(tables2)} tables to separate sheets in {excel_file}")

    remove_first_row_from_xlsx(excel_file, excel_file)

def parse_skyteam_timetable(cursor: sqlite3.Cursor, excel_file: str) -> None:
    """Parse the Skyteam_Timetable.xlsx file and insert data into the skyteam_timetable table.

    This function reads each relevant sheet in the Excel file (e.g., Sheet1_Table_*, Sheet2_Table_*),
    extracts FROM and TO information, and then parses the flight schedule rows.
    It skips sheets or rows that say 'Consult your travel agent for details' or are invalid.
    """
    try:
        xls = pd.ExcelFile(excel_file)

        for sheet_name in xls.sheet_names:
            if not (sheet_name.startswith('Sheet1_Table_') or sheet_name.startswith('Sheet2_Table_')):
                continue

            logger.info(f"Processing sheet: {sheet_name}")

            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)

            if df.empty:
                logger.warning(f"Sheet {sheet_name} is empty, skipping.")
                continue

            if len(df) < 3:
                logger.warning(f"Sheet {sheet_name} has insufficient rows, skipping.")
                continue

            from_row = df.iloc[0].dropna().tolist()
            if len(from_row) < 3 or from_row[0].strip() != 'FROM:':
                logger.warning(f"Invalid FROM row in sheet {sheet_name}, skipping.")
                continue

            if len(from_row) == 3:
                city_country = from_row[1].strip()
                if ',' in city_country:
                    from_city, from_country = [p.strip() for p in city_country.split(',', 1)]
                else:
                    from_city = city_country
                    from_country = ''
                from_code = from_row[2].strip()
            elif len(from_row) == 4:
                from_city = from_row[1].strip()
                from_country = from_row[2].strip()
                from_code = from_row[3].strip()
            else:
                logger.warning(f"Unexpected FROM row length in sheet {sheet_name}: {len(from_row)}, skipping.")
                continue

            to_row = df.iloc[1].dropna().tolist()
            if len(to_row) < 3 or to_row[0].strip() != 'TO:':
                logger.warning(f"Invalid TO row in sheet {sheet_name}, skipping.")
                continue

            if len(to_row) == 3:
                city_country = to_row[1].strip()
                if ',' in city_country:
                    to_city, to_country = [p.strip() for p in city_country.split(',', 1)]
                else:
                    to_city = city_country
                    to_country = ''
                to_code = to_row[2].strip()
            elif len(to_row) == 4:
                to_city = to_row[1].strip()
                to_country = to_row[2].strip()
                to_code = to_row[3].strip()
            else:
                logger.warning(f"Unexpected TO row length in sheet {sheet_name}: {len(to_row)}, skipping.")
                continue

            if len(df) == 4 and 'Consult your travel agent for details' in df.iloc[3].values:
                logger.info(f"No flight data in sheet {sheet_name}, skipping insertion.")
                continue

            header_row = df.iloc[2].tolist()
            if 'Validity' not in header_row or 'Days' not in header_row:
                logger.warning(f"Invalid header in sheet {sheet_name}, skipping.")
                continue

            data_df = df.iloc[3:].reset_index(drop=True)
            data_df.columns = header_row

            for _, row in data_df.iterrows():
                if pd.isna(row['Validity']) and pd.isna(row['Days']):
                    continue

                validity = str(row.get('Validity', '')).strip()
                days = str(row.get('Days', '')).strip()

                dep_time = ''
                if 'Dep\nTime' in row:
                    dep_time = str(row['Dep\nTime']).strip() if not pd.isna(row['Dep\nTime']) else ''
                elif 'Dep Time' in row:
                    dep_time = str(row['Dep Time']).strip() if not pd.isna(row['Dep Time']) else ''

                arr_time = ''
                if 'Arr\nTime' in row:
                    arr_time = str(row['Arr\nTime']).strip() if not pd.isna(row['Arr\nTime']) else ''
                elif 'Arr Time' in row:
                    arr_time = str(row['Arr Time']).strip() if not pd.isna(row['Arr Time']) else ''

                flight = str(row.get('Flight', '')).strip()
                aircraft = str(row.get('Aircraft', '')).strip()

                travel_time = ''
                if 'Travel\nTime' in row:
                    travel_time = str(row['Travel\nTime']).strip() if not pd.isna(row['Travel\nTime']) else ''
                elif 'Travel Time' in row:
                    travel_time = str(row['Travel Time']).strip() if not pd.isna(row['Travel Time']) else ''

                if validity and flight:
                    cursor.execute('''
                        INSERT INTO skyteam_timetable (
                            from_city, from_country, from_code,
                            to_city, to_country, to_code,
                            validity, days, dep_time, arr_time,
                            flight, aircraft, travel_time
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (from_city, from_country, from_code,
                          to_city, to_country, to_code,
                          validity, days, dep_time, arr_time,
                          flight, aircraft, travel_time))
                else:
                    logger.warning(f"Skipping invalid data row in {sheet_name}: {row}")

        logger.info("Skyteam timetable data successfully parsed and inserted.")

    except FileNotFoundError:
        logger.error(f"Excel file not found: {excel_file}")
        raise
    except pd.errors.EmptyDataError:
        logger.error(f"Empty data in Excel file: {excel_file}")
        raise
    except Exception as e:
        logger.error(f"Error parsing Skyteam timetable Excel file: {e}")
        raise

def parse_json_file(cursor: sqlite3.Cursor, json_file: str) -> None:
    """Parse the JSON file and insert data into the frequent_flyer_profiles and frequent_flyer_flights tables."""
    try:
        with open(json_file, 'r', encoding='utf-8') as file:
            data = json.load(file)
        profiles = data.get("Forum Profiles", [])
        for profile in profiles:
            nickname = profile.get("NickName", "")
            sex = profile.get("Sex", "")
            real_name = profile.get("Real Name", {})
            first_name = real_name.get("First Name", "") if real_name else ""
            last_name = real_name.get("Last Name", "") if real_name else ""
            travel_documents = json.dumps(profile.get("Travel Documents", []))
            loyalties = json.dumps(profile.get("Loyality Programm", []))
            cursor.execute('''
                INSERT OR REPLACE INTO frequent_flyer_profiles (Nick, Sex, FirstName, LastName, TravelDocuments, Loyalties)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (nickname, sex, first_name, last_name, travel_documents, loyalties))

            flights = profile.get("Registered Flights", [])
            for flight in flights:
                date = flight.get("Date", "")
                codeshare = 1 if flight.get("Codeshare", False) else 0
                flight_num = flight.get("Flight", "")
                dep = flight.get("Departure", {})
                dep_city = dep.get("City", "")
                dep_airport = dep.get("Airport", "")
                dep_country = dep.get("Country", "")
                arr = flight.get("Arrival", {})
                arr_city = arr.get("City", "")
                arr_airport = arr.get("Airport", "")
                arr_country = arr.get("Country", "")
                cursor.execute('''
                    INSERT INTO frequent_flyer_flights (
                        NickName, FlightDate, Flight, Codeshare,
                        DepartureCity, DepartureAirport, DepartureCountry,
                        ArrivalCity, ArrivalAirport, ArrivalCountry
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (nickname, date, flight_num, codeshare,
                      dep_city, dep_airport, dep_country,
                      arr_city, arr_airport, arr_country))
    except FileNotFoundError:
        logger.error(f"JSON file not found: {json_file}")
        raise
    except json.JSONDecodeError:
        logger.error(f"Invalid JSON format in file: {json_file}")
        raise
    except Exception as e:
        logger.error(f"Error parsing JSON file: {e}")
        raise

def parse_xls_files(cursor: sqlite3.Cursor, xls_dir: str) -> None:
    """Parse XLS files in the directory and insert data into the boarding_pass_xls table."""
    try:
        for file_path in Path(xls_dir).glob('*.xlsx'):
            xls = pd.ExcelFile(file_path)
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                if df.empty:
                    continue
                # Parse the data from the dataframe
                passenger_title = df.iloc[2, 0] if df.shape[0] > 2 and df.shape[1] > 0 else ''
                passenger_name = df.iloc[2, 1] if df.shape[1] > 1 else ''
                loyalty_str = df.iloc[2, 5] if df.shape[1] > 5 and pd.notna(df.iloc[2, 5]) else ''
                loyalty_program = ''
                loyalty_number = ''
                if loyalty_str.strip():
                    parts = loyalty_str.strip().split(' ', 1)
                    loyalty_program = parts[0] if len(parts) > 0 else ''
                    loyalty_number = parts[1] if len(parts) > 1 else ''
                fare_class = df.iloc[2, 7] if df.shape[1] > 7 and pd.notna(df.iloc[2, 7]) else ''
                flight_number = df.iloc[4, 0] if df.shape[0] > 4 else ''
                departure_city = df.iloc[4, 3] if df.shape[1] > 3 else ''
                arrival_city = df.iloc[4, 7] if df.shape[1] > 7 and pd.notna(df.iloc[4, 7]) else ''
                departure_airport = df.iloc[6, 3] if df.shape[1] > 3 else ''
                arrival_airport = df.iloc[6, 7] if df.shape[1] > 7 else ''
                flight_date = df.iloc[8, 0] if df.shape[0] > 8 else ''
                flight_time = df.iloc[8, 2] if df.shape[1] > 2 else ''
                pnr = df.iloc[12, 1] if df.shape[0] > 12 and df.shape[1] > 1 else ''
                eticket = df.iloc[12, 4] if df.shape[1] > 4 else ''

                cursor.execute('''
                    INSERT INTO boarding_pass_xls (
                        PassengerTitle, PassengerName, LoyaltyProgram, LoyaltyNumber, FareClass,
                        FlightNumber, DepartureCity, ArrivalCity, DepartureAirport, ArrivalAirport,
                        FlightDate, FlightTime, PNR, ETicket
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (passenger_title, passenger_name, loyalty_program, loyalty_number, fare_class,
                      flight_number, departure_city, arrival_city, departure_airport, arrival_airport,
                      flight_date, flight_time, pnr, eticket))
    except Exception as e:
        logger.error(f"Error parsing XLS files: {e}")
        raise

def main():
    """Main function to orchestrate database creation and file parsing."""
    try:
        for file in [CSV_FILE, TAB_FILE, XML_FILE, YAML_FILE, PDF_FILE, JSON_FILE]:
            if not Path(file).is_file():
                logger.error(f"File not found: {file}")
                return
        if not Path(XLS_DIR).exists():
            logger.error(f"Directory not found: {XLS_DIR}")
            return

        conn, cursor = create_database_connection(DB_FILE)

        # Create tables
        create_boarding_data_table(cursor) #csv
        create_sirena_data_table(cursor) #tab
        create_pointz_aggregator_table(cursor) #xml
        create_skyteam_data_table(cursor) #yaml
        create_skyteam_timetable_table(cursor) #pdf
        create_frequent_flyer_profiles_table(cursor) #json
        create_frequent_flyer_flights_table(cursor) #json
        create_boarding_pass_xls_table(cursor) #folder with xls

        # Process files based on flags
        if PROCESS_CSV:
            if CLEAR_CSV:
                logger.info("Clearing boarding_data table")
                cursor.execute('DELETE FROM boarding_data')
            logger.info(f"Processing CSV file: {CSV_FILE}")
            parse_csv_file(cursor, CSV_FILE)
        if PROCESS_TAB:
            if CLEAR_TAB:
                logger.info("Clearing sirena_data table")
                cursor.execute('DELETE FROM sirena_data')
            logger.info(f"Processing TAB file: {TAB_FILE}")
            parse_tab_file(cursor, TAB_FILE)
        if PROCESS_XML:
            if CLEAR_XML:
                logger.info("Clearing pointz_aggregator_data table")
                cursor.execute('DELETE FROM pointz_aggregator_data')
            logger.info(f"Processing XML file: {XML_FILE}")
            parse_xml_file(cursor, XML_FILE)
        if PROCESS_YAML:
            if CLEAR_YAML:
                logger.info("Clearing skyteam_data table")
                cursor.execute('DELETE FROM skyteam_data')
            logger.info(f"Processing YAML file: {YAML_FILE}")
            parse_yaml_file(cursor, YAML_FILE)
        if PROCESS_PDF:
            if CLEAR_PDF:
                logger.info("Clearing skyteam_timetable table")
                cursor.execute('DELETE FROM skyteam_timetable')
            logger.info(f"Processing PDF file: {PDF_FILE}")
            process_pdf_to_excel(PDF_FILE, "data/Skyteam_Timetable.xlsx", 3)
            #parse_skyteam_timetable(cursor, "data/Skyteam_Timetable.xlsx")
        if PROCESS_JSON:
            if CLEAR_JSON:
                logger.info("Clearing frequent_flyer_profiles table")
                cursor.execute('DELETE FROM frequent_flyer_profiles')
                logger.info("Clearing frequent_flyer_flights table")
                cursor.execute('DELETE FROM frequent_flyer_flights')
            logger.info(f"Processing JSON file: {JSON_FILE}")
            parse_json_file(cursor, JSON_FILE)
        if PROCESS_XLS:
            if CLEAR_XLS:
                logger.info("Clearing boarding_pass_xls table")
                cursor.execute('DELETE FROM boarding_pass_xls')
            logger.info(f"Processing XLS files in directory: {XLS_DIR}")
            parse_xls_files(cursor, XLS_DIR)

        # Commit changes and close connection
        conn.commit()
        logger.info("Data successfully inserted into the database.")
    except Exception as e:
        logger.error(f"An error occurred: {e}")
    finally:
        if 'conn' in locals():
            conn.close()

if __name__ == "__main__":
    main()